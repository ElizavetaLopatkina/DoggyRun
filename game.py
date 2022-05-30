import openpyxl
import pygame
from sys import exit
from random import randint
import os.path
import datetime

# excel database
xlsx_filepath = "./game.xlsx"

if not os.path.isfile(xlsx_filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Date'
    ws['B1'] = 'Score'
    wb.save(xlsx_filepath)


def get_free_row():
    fn = xlsx_filepath
    wb = openpyxl.load_workbook(fn)
    ws = wb.active
    global current_row
    for row in range(2, 10):
        if ws[f'B{row}'].value is None:
            current_row = row
            break

get_free_row()


#score
def display_score():


   current_time = int(pygame.time.get_ticks() / 1000) -start_time
   score_surf = test_font.render(f'Score: {current_time}',False,(64,64,64))
   score_rect = score_surf.get_rect(center = (600, 30))
   screen.blit(score_surf,score_rect)
   return current_time


#speed obtacle

def obstacle_movement(obstacle_list):
    if obstacle_list:
        for obstacle_rect in obstacle_list:
            obstacle_rect.x -= 7


            screen.blit(poop_surf, obstacle_rect)
        obstacle_list = [obstacle for obstacle in obstacle_list if obstacle.x > -100 ]

        return obstacle_list
    else: return []


def collision(corgi1,obstacles):
    if obstacles:
        for obstacles_rect in obstacles:
            if corgi1.colliderect(obstacles_rect): return False
    return True


def corgi1_animation():
    global corgi1_surf,corgi1_index

    if corgi1_rect.bottom < 500:
        corgi1_surf = corgi1_jump
    else:
        corgi1_index += 0.1
        if corgi1_index >= len(corgi1_walk):corgi1_index = 0
        corgi1_surf = corgi1_walk [int(corgi1_index)]



# Функция для паузы
def pause ():
    paused = True
    while paused:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit() #Закрытие игры
                exit ()

        pause_surf = test_font.render ('Pause', False, (64,64,64))
        pause_rect = pause_surf.get_rect (center = (1200 / 2, 800 / 2))
        screen.blit (pause_surf, pause_rect)


        keys = pygame.key.get_pressed ()
        if keys [pygame.K_RETURN] or keys [pygame.K_SPACE]:
            paused = False

        pygame.display.update ()
        clock.tick (15)




pygame.init()

screen = pygame.display.set_mode((1200, 700))
pygame.display.set_caption('Game')
clock = pygame.time.Clock()
test_font = pygame.font.Font('font/BUD_Pixel.ttf', 50)
game_active = False
start_time = 0
score = 0



background_surface = pygame.image.load('graphics/background.png').convert()

score_surf = test_font.render('Doggy run ', False, 'hotpink')
score_rect = score_surf.get_rect(center=(600, 90))

# преграды

poop_surf = pygame.image.load('graphics/poop.png').convert_alpha()
poop_rect = poop_surf.get_rect(bottomright=(600, 600))




#коржики
obstacle_rect_list = []


corgi1_walk_1 = pygame.image.load('graphics/corgi1.png').convert_alpha()
corgi1_walk_2 = pygame.image.load('graphics/corgi2.png').convert_alpha()
corgi1_walk = [corgi1_walk_1,corgi1_walk_2]
corgi1_index = 0
corgi1_jump = pygame.image.load('graphics/corgi1_jump.png').convert_alpha()


corgi1_surf = corgi1_walk[corgi1_index]
corgi1_rect = corgi1_surf.get_rect(midbottom=(100, 620))
corgi1_gravity = 0

# вступительная картинка
corgi1_stand = pygame.image.load('graphics/corgi1.png').convert_alpha()
corgi1_stand = pygame.transform.scale(corgi1_stand,(470,370))
corgi1_stand_rect = corgi1_stand.get_rect(center = (550,240))

game_name = test_font.render('Doggy run', False,'saddlebrown')
game_name_rect = game_name.get_rect(center=(580,120))

game_message = test_font.render('Press space to play', False,'saddlebrown')
game_message_rect = game_message.get_rect(center  = (580,550))

#Timer

obstacle_timer = pygame.USEREVENT + 1
pygame.time.set_timer(obstacle_timer, 2000)

while True:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            pygame.quit()
            exit()

        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                pause()

        if game_active:

            if event.type == pygame.MOUSEBUTTONDOWN:
                if corgi1_rect.collidepoint(event.pos) and corgi1_rect.bottom >= 620:
                    corgi1_gravity = -30

            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_SPACE and corgi1_rect.bottom >= 620:
                     corgi1_gravity = -30
        else:
            if event.type == pygame.KEYDOWN and  event.key == pygame.K_SPACE:
                game_active = True

                start_time =int(pygame.time.get_ticks() / 1000)

        if event.type == obstacle_timer and game_active:
            if randint(0,2):
             obstacle_rect_list.append(poop_surf.get_rect(bottomright=(randint(1250, 1500), 620)))

    if game_active:
        screen.blit(background_surface,(0,0))
        screen.blit(score_surf, score_rect)
        score = display_score()




        #коржик
        corgi1_gravity += 1.1
        corgi1_rect.y += corgi1_gravity
        if corgi1_rect.bottom >= 620:
            corgi1_rect.bottom = 620
        corgi1_animation()
        screen.blit(corgi1_surf,corgi1_rect)

        #движение пупа

        obstacle_rect_list = obstacle_movement(obstacle_rect_list)


        #collision
        game_active = collision(corgi1_rect,obstacle_rect_list)


       #надпись

    else:
        screen.fill('Plum')
        screen.blit(corgi1_stand,corgi1_stand_rect)
        screen.blit(game_name, game_name_rect)
        screen.blit(game_message,game_message_rect)
        score_message = test_font.render(f'Your score :{score}', False, 'mediumvioletred')
        score_message_rect = score_message.get_rect(center =(580,650) )

        if score == 0:
            screen.blit(game_message, game_message_rect)
        else:
            screen.blit(score_message,score_message_rect)
            obstacle_rect_list.clear()
            fn = xlsx_filepath
            wb = openpyxl.load_workbook(fn)
            ws = wb.active
            ws[f'B{current_row}'] = score
            ws[f'A{current_row}'] = str(datetime.datetime.now())
            wb.save(fn)
            wb.close



    pygame.display.update()
    clock.tick(60)



